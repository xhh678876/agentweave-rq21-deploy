"""
Test skill: xlsx
Verify that the Agent correctly creates a multi-sheet financial comparison
workbook using openpyxl with all formulas, formatting, and structure.
"""

import os
import subprocess
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"

    @classmethod
    def setup_class(cls):
        """Ensure openpyxl is importable; run the build script if needed."""
        try:
            import openpyxl
        except ImportError:
            subprocess.run(
                ["pip", "install", "-e", "."],
                cwd=cls.REPO_DIR,
                capture_output=True,
                timeout=120,
            )

    # === File Path Checks ===

    def test_build_script_exists(self):
        """Verify scripts/build_financial_report.py was created"""
        path = os.path.join(self.REPO_DIR, "scripts/build_financial_report.py")
        assert os.path.exists(path), f"Build script not found at {path}"

    def test_build_script_is_valid_python(self):
        """Verify build_financial_report.py is syntactically valid"""
        import ast
        path = os.path.join(self.REPO_DIR, "scripts/build_financial_report.py")
        with open(path) as f:
            source = f.read()
        try:
            ast.parse(source)
        except SyntaxError as e:
            pytest.fail(f"build_financial_report.py has syntax errors: {e}")

    # === Functional Check: Generate Workbook ===

    def _generate_workbook(self):
        """Helper to run the build script and return the output path."""
        result = subprocess.run(
            ["python", "scripts/build_financial_report.py"],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=120,
        )
        assert result.returncode == 0, (
            f"Build script failed: {result.stderr[-1000:]}\n{result.stdout[-500:]}"
        )
        output_path = os.path.join(self.REPO_DIR, "output/financial_comparison.xlsx")
        assert os.path.exists(output_path), (
            f"Expected output file not found at {output_path}"
        )
        return output_path

    def test_script_produces_xlsx_file(self):
        """Verify running the build script produces the xlsx output file"""
        self._generate_workbook()

    def test_workbook_has_four_sheets(self):
        """Verify workbook contains exactly 4 sheets with correct names"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        expected_sheets = {"Assumptions", "Division Summary", "Consolidated P&L", "Dashboard"}
        actual_sheets = set(wb.sheetnames)
        assert expected_sheets == actual_sheets, (
            f"Expected sheets {expected_sheets}, got {actual_sheets}"
        )

    # === Semantic Checks ===

    def test_assumptions_sheet_has_growth_rates(self):
        """Verify Assumptions sheet contains growth rate values"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Assumptions"]
        # Check C2 has a numeric value (growth rate)
        c2_val = ws["C2"].value
        assert c2_val is not None, "Assumptions C2 (first growth rate) is empty"
        assert isinstance(c2_val, (int, float)), (
            f"Growth rate should be numeric, got {type(c2_val)}: {c2_val}"
        )

    def test_assumptions_sheet_has_cogs_and_tax(self):
        """Verify Assumptions sheet has COGS % and Tax Rate"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Assumptions"]
        # Look for COGS and Tax Rate in column B
        found_cogs = False
        found_tax = False
        for row in ws.iter_rows(min_col=2, max_col=2, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "cogs" in cell.value.lower():
                        found_cogs = True
                    if "tax" in cell.value.lower():
                        found_tax = True
        assert found_cogs, "Assumptions sheet missing COGS % label"
        assert found_tax, "Assumptions sheet missing Tax Rate label"

    def test_division_summary_has_formulas(self):
        """Verify Division Summary uses Excel formulas, not hardcoded values"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Division Summary"]
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        assert formula_count >= 10, (
            f"Division Summary should have many formulas, found only {formula_count}"
        )

    def test_division_summary_has_three_divisions(self):
        """Verify Division Summary lists Cloud Services, Enterprise Software, Professional Services"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Division Summary"]
        content = ""
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    content += cell + " "
        content_lower = content.lower()
        assert "cloud" in content_lower, "Missing 'Cloud Services' division"
        assert "enterprise" in content_lower, "Missing 'Enterprise Software' division"
        assert "professional" in content_lower, "Missing 'Professional Services' division"

    def test_consolidated_pl_has_formulas(self):
        """Verify Consolidated P&L uses Excel formulas for computed rows"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Consolidated P&L"]
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        assert formula_count >= 20, (
            f"Consolidated P&L should have many formulas, found only {formula_count}"
        )

    def test_consolidated_pl_has_required_rows(self):
        """Verify P&L has Revenue, COGS, Gross Profit, EBIT, Taxes, Net Income rows"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Consolidated P&L"]
        labels = set()
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    labels.add(cell.lower().strip())
        label_text = " ".join(labels)
        assert "revenue" in label_text, "P&L missing Revenue row"
        assert "cogs" in label_text or "cost" in label_text, "P&L missing COGS row"
        assert "gross" in label_text, "P&L missing Gross Profit row"
        assert "ebit" in label_text or "operating" in label_text, "P&L missing EBIT row"
        assert "tax" in label_text, "P&L missing Taxes row"
        assert "net" in label_text, "P&L missing Net Income row"

    def test_taxes_formula_uses_max_zero(self):
        """Verify Taxes formula uses MAX(..., 0) to handle negative EBIT"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Consolidated P&L"]
        found_max = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "MAX" in cell.value.upper():
                    found_max = True
                    break
        assert found_max, (
            "Taxes formula should use MAX(..., 0) to prevent negative tax on losses"
        )

    def test_dashboard_has_formulas(self):
        """Verify Dashboard sheet uses formulas for computed metrics"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Dashboard"]
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        assert formula_count >= 4, (
            f"Dashboard should have at least 4 formulas for key metrics, found {formula_count}"
        )

    def test_cross_sheet_references_exist(self):
        """Verify formulas contain cross-sheet references (e.g., Assumptions!)"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        cross_refs = 0
        for sheet_name in ["Division Summary", "Consolidated P&L", "Dashboard"]:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "!" in cell.value and cell.value.startswith("="):
                        cross_refs += 1
        assert cross_refs >= 5, (
            f"Expected multiple cross-sheet references, found only {cross_refs}"
        )

    def test_no_hardcoded_computed_values(self):
        """Verify Division Summary computed cells are formulas, not Python-calculated numbers"""
        import openpyxl
        wb_path = self._generate_workbook()
        wb = openpyxl.load_workbook(wb_path)
        ws = wb["Division Summary"]
        # Check that cells after the base revenue (columns beyond B for Q2+) are formulas
        non_formula_computed = 0
        for row in ws.iter_rows(min_row=3, max_row=20, min_col=3, max_col=10):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    # This could be a hardcoded value - count them
                    non_formula_computed += 1
        # Allow a few numeric cells (like base revenue), but not too many
        assert non_formula_computed <= 10, (
            f"Found {non_formula_computed} numeric (non-formula) cells in computed area. "
            "Computed values should be Excel formulas, not Python-calculated numbers."
        )
