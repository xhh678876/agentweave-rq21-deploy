"""
Test skill: xlsx
Verify that the Agent correctly builds a report generation engine for
openpyxl that creates multi-sheet workbooks with data tables, summary
calculations, conditional formatting, and charts.
"""

import os
import sys
import ast
import subprocess
import tempfile
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"

    # === File Path Checks ===

    def test_report_engine_file_exists(self):
        """Verify report_engine.py exists in the correct location"""
        path = os.path.join(self.REPO_DIR, "openpyxl/utils/report_engine.py")
        assert os.path.exists(path), f"report_engine.py not found at {path}"

    def test_report_engine_is_valid_python(self):
        """Verify report_engine.py is syntactically valid Python"""
        path = os.path.join(self.REPO_DIR, "openpyxl/utils/report_engine.py")
        with open(path) as f:
            content = f.read()
        try:
            ast.parse(content)
        except SyntaxError as e:
            pytest.fail(f"report_engine.py has syntax error: {e}")

    # === Semantic Checks ===

    def test_report_engine_defines_callable(self):
        """Verify report_engine.py defines a callable function or class"""
        path = os.path.join(self.REPO_DIR, "openpyxl/utils/report_engine.py")
        with open(path) as f:
            tree = ast.parse(f.read())

        functions = [n.name for n in ast.walk(tree) if isinstance(n, ast.FunctionDef)]
        classes = [n.name for n in ast.walk(tree) if isinstance(n, ast.ClassDef)]

        has_entry_point = (
            any("report" in f.lower() or "generate" in f.lower() or "create" in f.lower()
                for f in functions)
            or any("report" in c.lower() or "engine" in c.lower() for c in classes)
        )
        assert has_entry_point, (
            f"report_engine.py should define a report generation callable. "
            f"Functions: {functions}, Classes: {classes}"
        )

    def test_report_engine_creates_multiple_sheets(self):
        """Verify report_engine.py references multiple sheet creation"""
        path = os.path.join(self.REPO_DIR, "openpyxl/utils/report_engine.py")
        with open(path) as f:
            content = f.read().lower()

        sheet_keywords = {
            "data_sheet": "data" in content and "sheet" in content,
            "summary_sheet": "summary" in content,
            "chart_sheet": "chart" in content,
        }
        found = [k for k, v in sheet_keywords.items() if v]
        assert len(found) >= 3, (
            f"Report engine should create data, summary, and chart sheets. "
            f"Found references to: {found}"
        )

    def test_report_engine_uses_conditional_formatting(self):
        """Verify report_engine.py applies conditional formatting"""
        path = os.path.join(self.REPO_DIR, "openpyxl/utils/report_engine.py")
        with open(path) as f:
            content = f.read()

        cf_indicators = [
            "ConditionalFormatting", "conditional_formatting",
            "CellIsRule", "FormulaRule", "ColorScaleRule",
            "DataBarRule", "IconSetRule",
        ]
        found = [ind for ind in cf_indicators if ind in content]
        assert len(found) >= 1, (
            "Report engine should apply conditional formatting. "
            f"None of {cf_indicators} found in source."
        )

    def test_report_engine_creates_charts(self):
        """Verify report_engine.py creates at least one chart type"""
        path = os.path.join(self.REPO_DIR, "openpyxl/utils/report_engine.py")
        with open(path) as f:
            content = f.read()

        chart_types = [
            "BarChart", "LineChart", "PieChart",
            "AreaChart", "ScatterChart", "chart",
        ]
        found = [ct for ct in chart_types if ct in content]
        assert len(found) >= 1, (
            "Report engine should create at least one chart. "
            f"None of {chart_types} found in source."
        )

    def test_report_engine_applies_styling(self):
        """Verify report_engine.py applies header styling and formatting"""
        path = os.path.join(self.REPO_DIR, "openpyxl/utils/report_engine.py")
        with open(path) as f:
            content = f.read()

        style_indicators = [
            "Font", "bold", "Alignment", "PatternFill",
            "Border", "column_dimensions", "number_format",
            "width",
        ]
        found = [ind for ind in style_indicators if ind in content]
        assert len(found) >= 2, (
            "Report engine should apply styling (bold headers, column widths, etc.). "
            f"Found: {found}. Expected at least 2 of: {style_indicators}"
        )

    # === Functional Checks ===

    def test_report_engine_is_importable(self):
        """Verify report_engine module can be imported"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.utils import report_engine
            assert report_engine is not None
        except ImportError as e:
            pytest.fail(f"Cannot import report_engine: {e}")

    def test_report_engine_generates_workbook(self):
        """Verify report engine generates a valid .xlsx workbook from sample data"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.utils import report_engine
        except ImportError:
            pytest.skip("Cannot import report_engine")

        # Find the main callable
        generate_func = None
        for name in dir(report_engine):
            obj = getattr(report_engine, name)
            if callable(obj) and (
                "report" in name.lower() or "generate" in name.lower()
                or "create" in name.lower()
            ):
                generate_func = obj
                break

        if generate_func is None:
            # Try if there's a class with a generate method
            for name in dir(report_engine):
                obj = getattr(report_engine, name)
                if isinstance(obj, type) and (
                    "report" in name.lower() or "engine" in name.lower()
                ):
                    try:
                        instance = obj()
                        for method_name in ["generate", "create", "build", "run"]:
                            if hasattr(instance, method_name):
                                generate_func = getattr(instance, method_name)
                                break
                    except TypeError:
                        pass
                    break

        if generate_func is None:
            pytest.fail(
                f"No report generation function found. Available: {dir(report_engine)}"
            )

        # Sample sales data
        sample_data = [
            {"date": "2024-01-01", "product": "Widget A", "quantity": 10, "revenue": 100.0},
            {"date": "2024-01-02", "product": "Widget B", "quantity": 5, "revenue": 75.0},
            {"date": "2024-01-03", "product": "Widget A", "quantity": 8, "revenue": 80.0},
            {"date": "2024-01-04", "product": "Widget C", "quantity": 15, "revenue": 225.0},
            {"date": "2024-01-05", "product": "Widget B", "quantity": 3, "revenue": 45.0},
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            result = generate_func(sample_data, output_path)
            # Result might be the path or a Workbook object
            if result is None:
                # Function might save to path directly
                assert os.path.exists(output_path), (
                    "Report engine should generate an xlsx file"
                )
            else:
                # Might return a Workbook to be saved
                try:
                    from openpyxl import Workbook
                    if isinstance(result, Workbook):
                        result.save(output_path)
                except Exception:
                    pass

            assert os.path.exists(output_path), (
                "Report engine should produce an .xlsx file"
            )
            assert os.path.getsize(output_path) > 100, (
                f"Generated xlsx is too small ({os.path.getsize(output_path)} bytes)"
            )
        except TypeError:
            pytest.skip("Report engine function signature not compatible with test input")
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_generated_workbook_has_multiple_sheets(self):
        """Verify generated workbook contains data, summary, and chart sheets"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.utils import report_engine
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Cannot import required modules")

        generate_func = None
        for name in dir(report_engine):
            obj = getattr(report_engine, name)
            if callable(obj) and (
                "report" in name.lower() or "generate" in name.lower()
                or "create" in name.lower()
            ):
                generate_func = obj
                break

        if generate_func is None:
            for name in dir(report_engine):
                obj = getattr(report_engine, name)
                if isinstance(obj, type):
                    try:
                        instance = obj()
                        for method_name in ["generate", "create", "build", "run"]:
                            if hasattr(instance, method_name):
                                generate_func = getattr(instance, method_name)
                                break
                    except TypeError:
                        pass

        if generate_func is None:
            pytest.skip("No report generation function found")

        sample_data = [
            {"date": "2024-01-01", "product": "A", "quantity": 10, "revenue": 100.0},
            {"date": "2024-01-02", "product": "B", "quantity": 5, "revenue": 75.0},
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            result = generate_func(sample_data, output_path)
            if result is not None:
                try:
                    from openpyxl import Workbook
                    if isinstance(result, Workbook):
                        result.save(output_path)
                except Exception:
                    pass

            if not os.path.exists(output_path) or os.path.getsize(output_path) < 100:
                pytest.skip("Workbook was not generated successfully")

            wb = load_workbook(output_path)
            sheet_names = wb.sheetnames
            assert len(sheet_names) >= 3, (
                f"Workbook should have at least 3 sheets (data, summary, chart). "
                f"Found {len(sheet_names)}: {sheet_names}"
            )
        except TypeError:
            pytest.skip("Report engine function signature not compatible")
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_data_sheet_has_headers_and_rows(self):
        """Verify data sheet contains header row and data rows"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.utils import report_engine
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Cannot import required modules")

        generate_func = None
        for name in dir(report_engine):
            obj = getattr(report_engine, name)
            if callable(obj) and (
                "report" in name.lower() or "generate" in name.lower()
                or "create" in name.lower()
            ):
                generate_func = obj
                break

        if generate_func is None:
            pytest.skip("No report generation function found")

        sample_data = [
            {"date": "2024-01-01", "product": "A", "quantity": 10, "revenue": 100.0},
            {"date": "2024-01-02", "product": "B", "quantity": 5, "revenue": 75.0},
            {"date": "2024-01-03", "product": "C", "quantity": 8, "revenue": 120.0},
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            result = generate_func(sample_data, output_path)
            if result is not None:
                try:
                    from openpyxl import Workbook
                    if isinstance(result, Workbook):
                        result.save(output_path)
                except Exception:
                    pass

            if not os.path.exists(output_path) or os.path.getsize(output_path) < 100:
                pytest.skip("Workbook not generated")

            wb = load_workbook(output_path)
            # First sheet is typically the data sheet
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
            assert len(rows) >= 4, (
                f"Data sheet should have header + {len(sample_data)} rows, "
                f"found {len(rows)} rows"
            )

            # Check header row contains expected field names
            header = [str(cell).lower() if cell else "" for cell in rows[0]]
            expected_fields = ["date", "product", "quantity", "revenue"]
            for field in expected_fields:
                found = any(field in h for h in header)
                assert found, (
                    f"Header row missing '{field}'. Headers: {rows[0]}"
                )
        except TypeError:
            pytest.skip("Report engine function signature not compatible")
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_summary_sheet_has_aggregate_values(self):
        """Verify summary sheet contains aggregate calculations"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.utils import report_engine
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Cannot import required modules")

        generate_func = None
        for name in dir(report_engine):
            obj = getattr(report_engine, name)
            if callable(obj) and (
                "report" in name.lower() or "generate" in name.lower()
                or "create" in name.lower()
            ):
                generate_func = obj
                break

        if generate_func is None:
            pytest.skip("No report generation function found")

        sample_data = [
            {"date": "2024-01-01", "product": "A", "quantity": 10, "revenue": 100.0},
            {"date": "2024-01-02", "product": "B", "quantity": 20, "revenue": 200.0},
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name

        try:
            result = generate_func(sample_data, output_path)
            if result is not None:
                try:
                    from openpyxl import Workbook
                    if isinstance(result, Workbook):
                        result.save(output_path)
                except Exception:
                    pass

            if not os.path.exists(output_path) or os.path.getsize(output_path) < 100:
                pytest.skip("Workbook not generated")

            wb = load_workbook(output_path)
            if len(wb.worksheets) < 2:
                pytest.fail("Workbook should have at least 2 sheets (data + summary)")

            summary_ws = wb.worksheets[1]
            all_values = []
            for row in summary_ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        all_values.append(cell)

            # Summary should have some content
            assert len(all_values) > 0, "Summary sheet should not be empty"

            # Check for aggregate-related labels or values
            str_values = [str(v).lower() for v in all_values]
            aggregate_keywords = ["total", "average", "avg", "count", "sum", "mean"]
            found = [kw for kw in aggregate_keywords if any(kw in sv for sv in str_values)]
            assert len(found) >= 1, (
                f"Summary sheet should contain aggregate labels. "
                f"Values found: {all_values[:20]}"
            )
        except TypeError:
            pytest.skip("Report engine function signature not compatible")
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
