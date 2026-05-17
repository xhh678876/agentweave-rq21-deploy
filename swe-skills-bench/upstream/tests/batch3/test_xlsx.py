"""
Test skill: xlsx
Verify that the Agent correctly implements a multi-sheet financial comparison
report generator using openpyxl.
"""

import os
import sys
import ast
import csv
import subprocess
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"

    # === File Path Checks ===

    def test_financial_report_script_exists(self):
        """Verify the financial report generator script exists"""
        path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        assert os.path.exists(path), f"financial_report.py not found at {path}"
        with open(path) as f:
            ast.parse(f.read())

    def test_sample_csv_exists(self):
        """Verify the sample revenue CSV file exists and is valid"""
        path = os.path.join(self.REPO_DIR, "openpyxl/sample/sample_revenue.csv")
        assert os.path.exists(path), f"sample_revenue.csv not found at {path}"
        with open(path, newline="") as f:
            reader = csv.reader(f)
            headers = next(reader)
        assert len(headers) >= 5, \
            f"CSV should have at least 5 columns, got {len(headers)}: {headers}"

    def test_test_file_exists(self):
        """Verify the agent's test file exists"""
        path = os.path.join(self.REPO_DIR, "openpyxl/tests/test_financial_report.py")
        assert os.path.exists(path), f"test_financial_report.py not found at {path}"

    # === Semantic Checks ===

    def test_csv_has_required_columns(self):
        """Verify sample CSV has all required columns"""
        path = os.path.join(self.REPO_DIR, "openpyxl/sample/sample_revenue.csv")
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
        required = ["Year", "Quarter", "BusinessUnit", "Revenue", "COGS", "OperatingExpenses"]
        for col in required:
            found = any(col.lower() == h.lower().strip() for h in headers)
            assert found, \
                f"CSV missing required column '{col}'. Found: {headers}"

    def test_csv_has_multiple_years_and_units(self):
        """Verify CSV covers at least 2 years and 3 business units"""
        path = os.path.join(self.REPO_DIR, "openpyxl/sample/sample_revenue.csv")
        years = set()
        units = set()
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Handle case-insensitive keys
                year_key = next((k for k in row.keys() if k.lower().strip() == "year"), None)
                unit_key = next((k for k in row.keys() if k.lower().strip() == "businessunit"), None)
                if year_key:
                    years.add(row[year_key].strip())
                if unit_key:
                    units.add(row[unit_key].strip())
        assert len(years) >= 2, f"CSV should have at least 2 years, found {years}"
        assert len(units) >= 3, f"CSV should have at least 3 business units, found {units}"

    def test_script_uses_openpyxl_imports(self):
        """Verify the script imports and uses openpyxl features"""
        path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        with open(path) as f:
            content = f.read()
        assert "openpyxl" in content, "Script should import openpyxl"
        # Check for key features
        features = ["Workbook", "chart", "formatting", "conditional"]
        found = sum(1 for feat in features if feat.lower() in content.lower())
        assert found >= 2, \
            f"Script should use openpyxl features (Workbook, chart, formatting). Found {found}"

    def test_script_creates_three_sheets(self):
        """Verify the script references three expected sheet names"""
        path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        with open(path) as f:
            content = f.read()
        expected_sheets = ["Raw Data", "Summary", "YoY Comparison"]
        for sheet in expected_sheets:
            assert sheet in content, \
                f"Script should create sheet named '{sheet}'"

    def test_script_uses_formulas_not_hardcoded(self):
        """Verify the script uses Excel formulas (SUMIFS, SUM, IF) not hardcoded values"""
        path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        with open(path) as f:
            content = f.read()
        formula_keywords = ["SUMIFS", "SUM(", "IF(", "IFERROR", "ISBLANK"]
        found = sum(1 for kw in formula_keywords if kw in content)
        assert found >= 2, \
            f"Script should use Excel formulas (SUMIFS, SUM, IF). Found {found} formula keywords"

    # === Functional Checks ===

    def _install_openpyxl(self):
        """Helper: install openpyxl"""
        result = subprocess.run(
            ["pip", "install", "-e", "."],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120
        )
        if result.returncode != 0:
            pytest.skip(f"Failed to install openpyxl: {result.stderr[:500]}")

    def test_script_runs_successfully(self):
        """Verify the financial report script runs without errors"""
        self._install_openpyxl()
        script_path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        result = subprocess.run(
            ["python", script_path],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120
        )
        assert result.returncode == 0, \
            f"Script failed:\n{result.stdout[:1000]}\n{result.stderr[:1000]}"

    def test_generated_workbook_has_three_sheets(self):
        """Verify the generated workbook has exactly 3 sheets with correct names"""
        self._install_openpyxl()
        # Run the script first
        script_path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        subprocess.run(
            ["python", script_path],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120
        )
        # Find the generated xlsx file
        test_code = '''
import os, glob, sys
sys.path.insert(0, ".")
from openpyxl import load_workbook

# Look for generated xlsx files
patterns = [
    "openpyxl/sample/*.xlsx",
    "*.xlsx",
    "output/*.xlsx",
    "openpyxl/sample/output/*.xlsx",
]
xlsx_files = []
for pattern in patterns:
    xlsx_files.extend(glob.glob(pattern))

if not xlsx_files:
    print("SKIP: No xlsx file found")
    sys.exit(0)

wb = load_workbook(xlsx_files[0])
sheet_names = wb.sheetnames
print(f"SHEETS:{sheet_names}")

expected = ["Raw Data", "Summary", "YoY Comparison"]
for name in expected:
    if name not in sheet_names:
        print(f"FAIL: Missing sheet '{name}'. Found: {sheet_names}")
        sys.exit(1)
print("PASS")
'''
        result = subprocess.run(
            ["python", "-c", test_code],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=60
        )
        if "SKIP" in result.stdout:
            pytest.skip("No xlsx file found after running script")
        assert "PASS" in result.stdout, \
            f"Sheet validation failed: {result.stdout}\n{result.stderr[:500]}"

    def test_generated_workbook_has_formulas(self):
        """Verify Summary sheet cells contain Excel formulas, not hardcoded values"""
        self._install_openpyxl()
        script_path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        subprocess.run(
            ["python", script_path],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120
        )
        test_code = '''
import os, glob, sys
sys.path.insert(0, ".")
from openpyxl import load_workbook

patterns = ["openpyxl/sample/*.xlsx", "*.xlsx", "output/*.xlsx"]
xlsx_files = []
for p in patterns:
    xlsx_files.extend(glob.glob(p))

if not xlsx_files:
    print("SKIP")
    sys.exit(0)

wb = load_workbook(xlsx_files[0])
if "Summary" not in wb.sheetnames:
    print("FAIL: No Summary sheet")
    sys.exit(1)

ws = wb["Summary"]
formula_count = 0
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formula_count += 1

if formula_count >= 3:
    print(f"PASS: Found {formula_count} formulas")
else:
    print(f"FAIL: Only {formula_count} formulas found. Summary should use Excel formulas, not hardcoded values")
'''
        result = subprocess.run(
            ["python", "-c", test_code],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=60
        )
        if "SKIP" in result.stdout:
            pytest.skip("No xlsx file found")
        assert "PASS" in result.stdout, \
            f"Formula check failed: {result.stdout}\n{result.stderr[:500]}"

    def test_generated_workbook_has_conditional_formatting(self):
        """Verify conditional formatting is applied to the workbook"""
        self._install_openpyxl()
        script_path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        subprocess.run(
            ["python", script_path],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120
        )
        test_code = '''
import glob, sys
sys.path.insert(0, ".")
from openpyxl import load_workbook

patterns = ["openpyxl/sample/*.xlsx", "*.xlsx", "output/*.xlsx"]
xlsx_files = []
for p in patterns:
    xlsx_files.extend(glob.glob(p))

if not xlsx_files:
    print("SKIP")
    sys.exit(0)

wb = load_workbook(xlsx_files[0])
total_cond_fmt = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    total_cond_fmt += len(ws.conditional_formatting)

if total_cond_fmt >= 1:
    print(f"PASS: Found {total_cond_fmt} conditional formatting rules")
else:
    print("FAIL: No conditional formatting rules found")
'''
        result = subprocess.run(
            ["python", "-c", test_code],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=60
        )
        if "SKIP" in result.stdout:
            pytest.skip("No xlsx file found")
        assert "PASS" in result.stdout, \
            f"Conditional formatting check failed: {result.stdout}\n{result.stderr[:500]}"

    def test_generated_workbook_has_chart(self):
        """Verify a chart is present on the Summary sheet"""
        self._install_openpyxl()
        script_path = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        subprocess.run(
            ["python", script_path],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120
        )
        test_code = '''
import glob, sys
sys.path.insert(0, ".")
from openpyxl import load_workbook

patterns = ["openpyxl/sample/*.xlsx", "*.xlsx", "output/*.xlsx"]
xlsx_files = []
for p in patterns:
    xlsx_files.extend(glob.glob(p))

if not xlsx_files:
    print("SKIP")
    sys.exit(0)

wb = load_workbook(xlsx_files[0])
if "Summary" not in wb.sheetnames:
    print("FAIL: No Summary sheet")
    sys.exit(1)

ws = wb["Summary"]
charts = ws._charts
if len(charts) >= 1:
    chart = charts[0]
    title = str(chart.title) if chart.title else ""
    print(f"PASS: Found {len(charts)} chart(s), title='{title}'")
else:
    print("FAIL: No charts found on Summary sheet")
'''
        result = subprocess.run(
            ["python", "-c", test_code],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=60
        )
        if "SKIP" in result.stdout:
            pytest.skip("No xlsx file found")
        assert "PASS" in result.stdout, \
            f"Chart check failed: {result.stdout}\n{result.stderr[:500]}"

    def test_script_validates_csv_columns(self):
        """Verify script raises ValueError for CSV with missing columns"""
        self._install_openpyxl()
        test_code = '''
import sys, tempfile, os
sys.path.insert(0, ".")

# Create a CSV with missing columns
with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
    f.write("Year,Quarter,Revenue\\n")
    f.write("2023,Q1,1000\\n")
    bad_csv = f.name

try:
    # Try importing and running with bad CSV
    try:
        from openpyxl.sample.financial_report import generate_report
        try:
            generate_report(bad_csv, "test_output.xlsx")
            print("FAIL: Should have raised ValueError for missing columns")
        except ValueError as e:
            print(f"PASS: {e}")
        except Exception as e:
            print(f"PASS_OTHER: Raised {type(e).__name__}: {e}")
    except ImportError:
        # Try running as script with bad CSV
        import subprocess
        result = subprocess.run(
            [sys.executable, "openpyxl/sample/financial_report.py", bad_csv],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode != 0:
            print("PASS: Script failed with bad CSV as expected")
        else:
            print("FAIL: Script should fail with bad CSV")
finally:
    os.unlink(bad_csv)
'''
        result = subprocess.run(
            ["python", "-c", test_code],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=60
        )
        assert "PASS" in result.stdout, \
            f"CSV validation test failed: {result.stdout}\n{result.stderr[:500]}"
