"""
Test skill: xlsx
Verify that the Agent creates a budget_model.py generating a multi-sheet Excel workbook
with formulas, formatting, and data validation using openpyxl.
"""

import os
import subprocess
import ast
import re
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"

    # === File Path Checks ===

    def test_budget_model_script_exists(self):
        """Verify budget_model.py exists"""
        path = os.path.join(self.REPO_DIR, "budget_model.py")
        assert os.path.exists(path), f"budget_model.py not found at {path}"

    def test_output_directory_exists(self):
        """Verify output directory exists or will be created"""
        # Script should create the output dir or it exists already
        path = os.path.join(self.REPO_DIR, "output")
        script_path = os.path.join(self.REPO_DIR, "budget_model.py")
        with open(script_path) as f:
            content = f.read()
        has_mkdir = "makedirs" in content or "mkdir" in content or "Path" in content
        assert os.path.isdir(path) or has_mkdir, (
            "output directory doesn't exist and script doesn't create it"
        )

    # === Semantic Checks ===

    def test_budget_model_imports_openpyxl(self):
        """Verify budget_model.py imports openpyxl"""
        path = os.path.join(self.REPO_DIR, "budget_model.py")
        with open(path) as f:
            source = f.read()
        assert "openpyxl" in source, "budget_model.py does not import openpyxl"

    def test_budget_model_creates_multiple_sheets(self):
        """Verify script creates Assumptions, Revenue, Expenses, Summary sheets"""
        path = os.path.join(self.REPO_DIR, "budget_model.py")
        with open(path) as f:
            source = f.read()
        source_lower = source.lower()
        expected_sheets = ["assumptions", "revenue", "expenses", "summary"]
        found = [s for s in expected_sheets if s in source_lower]
        assert len(found) >= 3, (
            f"Expected sheets {expected_sheets}, found {found}"
        )

    def test_budget_model_uses_excel_formulas(self):
        """Verify script writes Excel formulas (=SUM, =IF, etc.) not just static values"""
        path = os.path.join(self.REPO_DIR, "budget_model.py")
        with open(path) as f:
            source = f.read()
        formula_pattern = re.compile(r"['\"]=[A-Z]+\(")
        has_formulas = bool(formula_pattern.search(source))
        assert has_formulas, "No Excel formulas (=SUM, =IF, etc.) found in budget_model.py"

    def test_budget_model_applies_formatting(self):
        """Verify script applies cell formatting (fonts, fills, number formats)"""
        path = os.path.join(self.REPO_DIR, "budget_model.py")
        with open(path) as f:
            source = f.read()
        formatting_indicators = ["Font", "PatternFill", "Alignment", "Border", "number_format"]
        found = [fi for fi in formatting_indicators if fi in source]
        assert len(found) >= 2, (
            f"Insufficient formatting. Found: {found}, expected at least 2 of {formatting_indicators}"
        )

    def test_budget_model_uses_data_validation(self):
        """Verify script includes data validation"""
        path = os.path.join(self.REPO_DIR, "budget_model.py")
        with open(path) as f:
            source = f.read()
        has_validation = (
            "DataValidation" in source
            or "data_validation" in source
            or "validation" in source.lower()
        )
        assert has_validation, "No data validation found in budget_model.py"

    # === Functional Checks ===

    def test_budget_model_parses_without_errors(self):
        """Verify budget_model.py is valid Python"""
        path = os.path.join(self.REPO_DIR, "budget_model.py")
        with open(path) as f:
            source = f.read()
        try:
            ast.parse(source)
        except SyntaxError as e:
            pytest.fail(f"budget_model.py has syntax errors: {e}")

    def test_budget_model_runs_successfully(self):
        """Verify budget_model.py executes and produces output file"""
        result = subprocess.run(
            ["python", "budget_model.py"],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert result.returncode == 0, (
            f"budget_model.py execution failed:\n{result.stderr[:500]}"
        )

    def test_output_xlsx_file_exists(self):
        """Verify department_budget.xlsx is created in output directory"""
        path = os.path.join(self.REPO_DIR, "output/department_budget.xlsx")
        if not os.path.exists(path):
            # Run the script first
            subprocess.run(
                ["python", "budget_model.py"],
                cwd=self.REPO_DIR,
                capture_output=True,
                text=True,
                timeout=60,
            )
        assert os.path.exists(path), f"department_budget.xlsx not found at {path}"

    def test_output_xlsx_has_correct_sheets(self):
        """Verify the generated xlsx has all required sheets"""
        script = """
import sys
sys.path.insert(0, '.')
import openpyxl
wb = openpyxl.load_workbook('output/department_budget.xlsx')
sheets = wb.sheetnames
print('SHEETS:' + ','.join(sheets))
expected = {'Assumptions', 'Revenue', 'Expenses', 'Summary'}
found = set(sheets) & expected
missing = expected - set(sheets)
if len(found) >= 3:
    print('PASS')
else:
    print(f'FAIL:missing {missing}')
"""
        # Ensure the file is generated first
        subprocess.run(
            ["python", "budget_model.py"],
            cwd=self.REPO_DIR,
            capture_output=True,
            timeout=60,
        )
        result = subprocess.run(
            ["python", "-c", script],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"Sheet check script failed: {result.stderr}"
        assert "PASS" in result.stdout, f"Sheet check failed: {result.stdout}"

    def test_output_xlsx_has_formulas(self):
        """Verify the generated xlsx contains Excel formulas, not just values"""
        script = """
import sys
sys.path.insert(0, '.')
import openpyxl
wb = openpyxl.load_workbook('output/department_budget.xlsx')
formula_count = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
if formula_count > 0:
    print(f'PASS:{formula_count}')
else:
    print('FAIL:no formulas found')
"""
        subprocess.run(
            ["python", "budget_model.py"],
            cwd=self.REPO_DIR,
            capture_output=True,
            timeout=60,
        )
        result = subprocess.run(
            ["python", "-c", script],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"Formula check script failed: {result.stderr}"
        assert "PASS" in result.stdout, f"No formulas in xlsx: {result.stdout}"
