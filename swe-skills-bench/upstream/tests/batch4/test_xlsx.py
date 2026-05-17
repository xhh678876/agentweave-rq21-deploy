"""
Test skill: xlsx
Verify that the multi-sheet financial summary workbook script has been correctly
implemented using openpyxl, including four sheets (Assumptions, Income Statement,
DCF Valuation, Dashboard), Excel formulas, formatting, charts, and data validation.
"""

import os
import sys
import subprocess
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"
    SCRIPT_PATH = "examples/financial_summary.py"
    OUTPUT_PATH = "examples/financial_summary_output.xlsx"

    @classmethod
    def _ensure_openpyxl_installed(cls):
        """Ensure openpyxl is available for testing"""
        try:
            import openpyxl
        except ImportError:
            subprocess.run(
                ["pip", "install", "-e", "."],
                cwd=cls.REPO_DIR,
                capture_output=True, text=True, timeout=120
            )

    @classmethod
    def _run_script_if_needed(cls):
        """Run the financial summary script if output doesn't exist"""
        output_path = os.path.join(cls.REPO_DIR, cls.OUTPUT_PATH)
        if not os.path.exists(output_path):
            cls._ensure_openpyxl_installed()
            subprocess.run(
                ["python", cls.SCRIPT_PATH],
                cwd=cls.REPO_DIR,
                capture_output=True, text=True, timeout=120
            )

    # === File Path Checks ===

    def test_script_file_exists(self):
        """Verify that examples/financial_summary.py exists"""
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        assert os.path.exists(filepath), f"Script not found at {filepath}"

    def test_script_is_valid_python(self):
        """Verify the script is syntactically valid Python"""
        import ast
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        with open(filepath) as f:
            content = f.read()
        try:
            ast.parse(content)
        except SyntaxError as e:
            pytest.fail(f"Script has syntax error: {e}")

    # === Semantic Checks ===

    def test_script_uses_openpyxl(self):
        """Verify the script imports and uses openpyxl"""
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        with open(filepath) as f:
            content = f.read()
        assert "openpyxl" in content, "Script should import openpyxl"
        assert "Workbook" in content, "Script should use openpyxl.Workbook"

    def test_script_creates_four_sheets(self):
        """Verify the script creates four named sheets"""
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        with open(filepath) as f:
            content = f.read()
        expected_sheets = ["Assumptions", "Income Statement", "DCF Valuation", "Dashboard"]
        for sheet_name in expected_sheets:
            assert sheet_name in content, \
                f"Script should create sheet named '{sheet_name}'"

    def test_script_uses_formulas_not_constants(self):
        """Verify that the script writes Excel formulas (=...) not hardcoded values"""
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        with open(filepath) as f:
            content = f.read()
        # Should contain Excel formula references
        has_formulas = ("'=" in content or '"=' in content or "= '=" in content or
                        '= "=' in content or ".value = '=" in content or
                        ".value = f'=" in content)
        assert has_formulas, \
            "Script should write Excel formulas (starting with '=') not hardcoded values"

    def test_script_uses_cross_sheet_references(self):
        """Verify script uses cross-sheet formula references"""
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        with open(filepath) as f:
            content = f.read()
        # Cross-sheet references contain sheet name with '!'
        has_cross_ref = ("!" in content and ("Assumptions" in content or
                                              "Income" in content or
                                              "DCF" in content))
        assert has_cross_ref, \
            "Script should use cross-sheet formula references (e.g., Assumptions!B2)"

    def test_script_implements_conditional_formatting(self):
        """Verify script applies conditional formatting or number format for negatives"""
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        with open(filepath) as f:
            content = f.read()
        has_formatting = ("PatternFill" in content or "Font" in content or
                          "number_format" in content or "conditional" in content.lower())
        assert has_formatting, \
            "Script should apply formatting (PatternFill, Font, number_format)"

    def test_script_creates_chart(self):
        """Verify script creates a chart on the Dashboard sheet"""
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        with open(filepath) as f:
            content = f.read()
        has_chart = ("BarChart" in content or "Chart" in content or
                     "chart" in content.lower())
        assert has_chart, "Script should create a bar chart on the Dashboard sheet"

    def test_script_creates_data_validation(self):
        """Verify script creates data validation dropdown for Scenario cell"""
        filepath = os.path.join(self.REPO_DIR, self.SCRIPT_PATH)
        with open(filepath) as f:
            content = f.read()
        has_validation = ("DataValidation" in content or "data_validation" in content)
        assert has_validation, \
            "Script should create a DataValidation dropdown on the Dashboard"

    # === Functional Checks ===

    def test_script_runs_without_error(self):
        """Verify the script executes successfully and produces output file"""
        self._ensure_openpyxl_installed()
        result = subprocess.run(
            ["python", self.SCRIPT_PATH],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120
        )
        assert result.returncode == 0, \
            f"Script execution failed: {result.stderr[:1000]}"
        output_path = os.path.join(self.REPO_DIR, self.OUTPUT_PATH)
        assert os.path.exists(output_path), \
            f"Output file not created at {output_path}"

    def test_output_has_four_sheets(self):
        """Verify the output xlsx has exactly four sheets with correct names"""
        self._run_script_if_needed()
        self._ensure_openpyxl_installed()

        import openpyxl
        output_path = os.path.join(self.REPO_DIR, self.OUTPUT_PATH)
        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames
        expected = ["Assumptions", "Income Statement", "DCF Valuation", "Dashboard"]
        for name in expected:
            assert name in sheet_names, \
                f"Workbook missing sheet '{name}'. Found: {sheet_names}"
        wb.close()

    def test_assumptions_sheet_has_named_ranges(self):
        """Verify the Assumptions sheet defines named ranges for input cells"""
        self._run_script_if_needed()
        self._ensure_openpyxl_installed()

        import openpyxl
        output_path = os.path.join(self.REPO_DIR, self.OUTPUT_PATH)
        wb = openpyxl.load_workbook(output_path)
        defined_names = [dn.name for dn in wb.defined_names.definedName]
        assert len(defined_names) >= 3, \
            f"Workbook should have at least 3 named ranges for assumptions, found {len(defined_names)}: {defined_names}"
        wb.close()

    def test_income_statement_has_formulas(self):
        """Verify Income Statement cells contain Excel formulas, not static values"""
        self._run_script_if_needed()
        self._ensure_openpyxl_installed()

        import openpyxl
        output_path = os.path.join(self.REPO_DIR, self.OUTPUT_PATH)
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Income Statement"]

        formula_count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

        assert formula_count >= 10, \
            f"Income Statement should have at least 10 formula cells, found {formula_count}"
        wb.close()

    def test_dcf_sheet_has_formulas(self):
        """Verify DCF Valuation sheet contains formulas for discount factors and PV"""
        self._run_script_if_needed()
        self._ensure_openpyxl_installed()

        import openpyxl
        output_path = os.path.join(self.REPO_DIR, self.OUTPUT_PATH)
        wb = openpyxl.load_workbook(output_path)
        ws = wb["DCF Valuation"]

        formula_count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

        assert formula_count >= 5, \
            f"DCF Valuation should have at least 5 formula cells, found {formula_count}"
        wb.close()

    def test_dashboard_has_chart(self):
        """Verify Dashboard sheet contains at least one chart"""
        self._run_script_if_needed()
        self._ensure_openpyxl_installed()

        import openpyxl
        output_path = os.path.join(self.REPO_DIR, self.OUTPUT_PATH)
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Dashboard"]
        chart_count = len(ws._charts)
        assert chart_count >= 1, \
            f"Dashboard should have at least 1 chart, found {chart_count}"
        wb.close()

    def test_output_file_size_is_reasonable(self):
        """Verify the output xlsx file is non-trivial (not empty or corrupted)"""
        self._run_script_if_needed()
        output_path = os.path.join(self.REPO_DIR, self.OUTPUT_PATH)
        file_size = os.path.getsize(output_path)
        assert file_size > 5000, \
            f"Output file seems too small ({file_size} bytes), may be corrupted"
        assert file_size < 10_000_000, \
            f"Output file seems too large ({file_size} bytes), may be malformed"
