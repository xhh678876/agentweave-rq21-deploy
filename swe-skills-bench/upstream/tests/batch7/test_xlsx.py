"""
Test skill: xlsx
Verify that the Agent correctly creates a financial report generator module
using openpyxl with multi-sheet workbooks, formulas, charts, and formatting.
"""

import os
import re
import ast
import sys
import subprocess
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"

    # === File Path Checks ===

    def test_report_package_init_exists(self):
        """Verify report/__init__.py package init exists"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/__init__.py")
        assert os.path.isfile(fpath), f"report/__init__.py not found at {fpath}"

    def test_generator_module_exists(self):
        """Verify report/generator.py exists"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/generator.py")
        assert os.path.isfile(fpath), f"report/generator.py not found at {fpath}"

    def test_styles_module_exists(self):
        """Verify report/styles.py exists"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/styles.py")
        assert os.path.isfile(fpath), f"report/styles.py not found at {fpath}"

    def test_charts_module_exists(self):
        """Verify report/charts.py exists"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/charts.py")
        assert os.path.isfile(fpath), f"report/charts.py not found at {fpath}"

    # === Semantic Checks ===

    def test_generator_has_class_definition(self):
        """Verify generator.py defines FinancialReportGenerator class"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/generator.py")
        with open(fpath, "r") as f:
            tree = ast.parse(f.read())
        classes = [n.name for n in ast.walk(tree) if isinstance(n, ast.ClassDef)]
        assert "FinancialReportGenerator" in classes, (
            f"FinancialReportGenerator class not found. Found: {classes}"
        )

    def test_generator_has_generate_method(self):
        """Verify FinancialReportGenerator has generate and generate_bytes methods"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/generator.py")
        with open(fpath, "r") as f:
            content = f.read()
        has_generate = bool(re.search(r'def\s+generate\s*\(', content))
        has_generate_bytes = bool(re.search(r'def\s+generate_bytes\s*\(', content))
        assert has_generate, "FinancialReportGenerator should have a generate() method"
        assert has_generate_bytes, "FinancialReportGenerator should have a generate_bytes() method"

    def test_generator_creates_three_sheets(self):
        """Verify generator references all three required sheet names"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/generator.py")
        with open(fpath, "r") as f:
            content = f.read()
        assert "Income Statement" in content, "Generator should create 'Income Statement' sheet"
        assert "Balance Sheet" in content, "Generator should create 'Balance Sheet' sheet"
        assert "Dashboard" in content, "Generator should create 'Dashboard' sheet"

    def test_generator_uses_excel_formulas(self):
        """Verify generator uses Excel formulas (=SUM, cell references) not hardcoded values"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/generator.py")
        with open(fpath, "r") as f:
            content = f.read()
        has_sum_formula = bool(re.search(r'["\']?=SUM\(', content))
        has_cell_ref = bool(re.search(r'[A-Z]\d+', content))
        assert has_sum_formula, "Generator should use =SUM formulas"
        assert has_cell_ref, "Generator should use cell references (e.g., A1, B2)"

    def test_charts_module_creates_bar_line_pie(self):
        """Verify charts.py creates bar, line, and pie chart types"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/charts.py")
        with open(fpath, "r") as f:
            content = f.read()
        has_bar = bool(re.search(r'(BarChart|bar_chart|bar)', content, re.IGNORECASE))
        has_line = bool(re.search(r'(LineChart|line_chart|line)', content, re.IGNORECASE))
        has_pie = bool(re.search(r'(PieChart|pie_chart|pie)', content, re.IGNORECASE))
        assert has_bar, "charts.py should create bar charts"
        assert has_line, "charts.py should create line charts"
        assert has_pie, "charts.py should create pie charts"

    def test_styles_has_currency_format(self):
        """Verify styles.py defines currency formatting"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/styles.py")
        with open(fpath, "r") as f:
            content = f.read()
        has_currency = bool(re.search(r'(#,##0|currency|accounting|dollar|\$)', content, re.IGNORECASE))
        assert has_currency, "styles.py should define currency number formatting"

    def test_styles_has_conditional_formatting(self):
        """Verify styles.py defines conditional formatting rules"""
        fpath = os.path.join(self.REPO_DIR, "openpyxl/report/styles.py")
        with open(fpath, "r") as f:
            content = f.read()
        has_conditional = bool(re.search(r'(conditional|ConditionalFormatting|PatternFill|red|RED)', content, re.IGNORECASE))
        assert has_conditional, "styles.py should define conditional formatting rules"

    # === Functional Checks ===

    def test_import_financial_report_generator(self):
        """Verify FinancialReportGenerator can be imported"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.report import FinancialReportGenerator
            assert FinancialReportGenerator is not None
        except ImportError as e:
            # Try alternative import path
            try:
                from openpyxl.report.generator import FinancialReportGenerator
                assert FinancialReportGenerator is not None
            except ImportError:
                pytest.fail(f"Cannot import FinancialReportGenerator: {e}")

    def test_generator_rejects_invalid_data(self):
        """Verify generator raises ValueError for missing required data keys"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.report.generator import FinancialReportGenerator
        except ImportError:
            pytest.skip("Cannot import FinancialReportGenerator")

        with pytest.raises((ValueError, KeyError, TypeError)) as exc_info:
            gen = FinancialReportGenerator({})
            gen.generate("/tmp/test_invalid.xlsx")
        # Error should be descriptive
        assert exc_info.value is not None

    def test_generator_creates_valid_workbook(self):
        """Verify generator creates a valid xlsx file with three sheets"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.report.generator import FinancialReportGenerator
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Cannot import required modules")

        sample_data = {
            "revenues": {"Q1": 100000, "Q2": 120000, "Q3": 110000, "Q4": 130000},
            "cogs": {"Q1": 40000, "Q2": 48000, "Q3": 44000, "Q4": 52000},
            "expenses": {
                "salaries": {"Q1": 20000, "Q2": 20000, "Q3": 22000, "Q4": 22000},
                "rent": {"Q1": 5000, "Q2": 5000, "Q3": 5000, "Q4": 5000},
                "marketing": {"Q1": 8000, "Q2": 10000, "Q3": 9000, "Q4": 12000},
                "other": {"Q1": 2000, "Q2": 2500, "Q3": 2000, "Q4": 3000},
            },
            "interest_expense": {"Q1": 1000, "Q2": 1000, "Q3": 1000, "Q4": 1000},
            "assets": {
                "cash": 50000, "accounts_receivable": 30000,
                "inventory": 20000, "fixed_assets": 100000
            },
            "liabilities": {
                "accounts_payable": 15000, "short_term_debt": 10000,
                "long_term_debt": 75000
            },
            "equity": {
                "common_stock": 50000, "retained_earnings": 50000
            }
        }

        output_path = "/tmp/test_financial_report.xlsx"
        try:
            gen = FinancialReportGenerator(sample_data)
            gen.generate(output_path)
            assert os.path.isfile(output_path), "Generator did not create output file"

            wb = load_workbook(output_path)
            sheet_names = wb.sheetnames
            assert "Income Statement" in sheet_names, f"Missing 'Income Statement' sheet. Found: {sheet_names}"
            assert "Balance Sheet" in sheet_names, f"Missing 'Balance Sheet' sheet. Found: {sheet_names}"
            assert "Dashboard" in sheet_names, f"Missing 'Dashboard' sheet. Found: {sheet_names}"
        except (ValueError, KeyError, TypeError) as e:
            pytest.fail(f"Generator failed with valid data: {e}")
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)

    def test_income_statement_has_formulas(self):
        """Verify Income Statement sheet contains formula cells"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.report.generator import FinancialReportGenerator
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Cannot import required modules")

        sample_data = {
            "revenues": {"Q1": 100000, "Q2": 120000},
            "cogs": {"Q1": 40000, "Q2": 48000},
            "expenses": {
                "salaries": {"Q1": 20000, "Q2": 20000},
                "rent": {"Q1": 5000, "Q2": 5000},
                "marketing": {"Q1": 8000, "Q2": 10000},
                "other": {"Q1": 2000, "Q2": 2500},
            },
            "interest_expense": {"Q1": 1000, "Q2": 1000},
            "assets": {"cash": 50000, "accounts_receivable": 30000,
                       "inventory": 20000, "fixed_assets": 100000},
            "liabilities": {"accounts_payable": 15000, "short_term_debt": 10000,
                           "long_term_debt": 75000},
            "equity": {"common_stock": 50000, "retained_earnings": 50000}
        }

        output_path = "/tmp/test_formulas.xlsx"
        try:
            gen = FinancialReportGenerator(sample_data)
            gen.generate(output_path)
            wb = load_workbook(output_path)
            ws = wb["Income Statement"]

            # Check for formula cells (cells starting with '=')
            formula_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
            assert formula_count >= 3, (
                f"Income Statement should have at least 3 formula cells, found {formula_count}"
            )
        except (ValueError, KeyError) as e:
            pytest.fail(f"Failed to generate report: {e}")
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)

    def test_dashboard_has_charts(self):
        """Verify Dashboard sheet contains chart objects"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.report.generator import FinancialReportGenerator
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("Cannot import required modules")

        sample_data = {
            "revenues": {"Q1": 100000, "Q2": 120000},
            "cogs": {"Q1": 40000, "Q2": 48000},
            "expenses": {
                "salaries": {"Q1": 20000, "Q2": 20000},
                "rent": {"Q1": 5000, "Q2": 5000},
                "marketing": {"Q1": 8000, "Q2": 10000},
                "other": {"Q1": 2000, "Q2": 2500},
            },
            "interest_expense": {"Q1": 1000, "Q2": 1000},
            "assets": {"cash": 50000, "accounts_receivable": 30000,
                       "inventory": 20000, "fixed_assets": 100000},
            "liabilities": {"accounts_payable": 15000, "short_term_debt": 10000,
                           "long_term_debt": 75000},
            "equity": {"common_stock": 50000, "retained_earnings": 50000}
        }

        output_path = "/tmp/test_charts.xlsx"
        try:
            gen = FinancialReportGenerator(sample_data)
            gen.generate(output_path)
            wb = load_workbook(output_path)
            dash = wb["Dashboard"]
            charts = dash._charts
            assert len(charts) >= 3, (
                f"Dashboard should have at least 3 charts (bar, line, pie), found {len(charts)}"
            )
        except (ValueError, KeyError) as e:
            pytest.fail(f"Failed to generate report: {e}")
        finally:
            if os.path.exists(output_path):
                os.remove(output_path)

    def test_generate_bytes_returns_bytes(self):
        """Verify generate_bytes() returns bytes that can be loaded as a workbook"""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl.report.generator import FinancialReportGenerator
            from openpyxl import load_workbook
            import io
        except ImportError:
            pytest.skip("Cannot import required modules")

        sample_data = {
            "revenues": {"Q1": 100000, "Q2": 120000},
            "cogs": {"Q1": 40000, "Q2": 48000},
            "expenses": {
                "salaries": {"Q1": 20000, "Q2": 20000},
                "rent": {"Q1": 5000, "Q2": 5000},
                "marketing": {"Q1": 8000, "Q2": 10000},
                "other": {"Q1": 2000, "Q2": 2500},
            },
            "interest_expense": {"Q1": 1000, "Q2": 1000},
            "assets": {"cash": 50000, "accounts_receivable": 30000,
                       "inventory": 20000, "fixed_assets": 100000},
            "liabilities": {"accounts_payable": 15000, "short_term_debt": 10000,
                           "long_term_debt": 75000},
            "equity": {"common_stock": 50000, "retained_earnings": 50000}
        }

        try:
            gen = FinancialReportGenerator(sample_data)
            data = gen.generate_bytes()
            assert isinstance(data, bytes), f"generate_bytes should return bytes, got {type(data)}"
            assert len(data) > 0, "generate_bytes should return non-empty bytes"
            # Verify it's a valid xlsx
            wb = load_workbook(io.BytesIO(data))
            assert len(wb.sheetnames) == 3, f"Workbook should have 3 sheets, got {wb.sheetnames}"
        except (ValueError, KeyError) as e:
            pytest.fail(f"generate_bytes failed: {e}")
