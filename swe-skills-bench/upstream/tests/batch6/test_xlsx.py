"""
Test skill: xlsx
Verify that the Agent correctly builds a financial reporting workbook
from raw transaction data using openpyxl with 4 sheets, formulas, and charts.
"""

import os
import re
import csv
import subprocess
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"

    # === File Path Checks ===

    def test_build_script_exists(self):
        """Verify that the build script exists"""
        path = os.path.join(self.REPO_DIR, "scripts/build_financial_report.py")
        assert os.path.exists(path), f"build_financial_report.py not found at {path}"

    def test_transactions_csv_exists(self):
        """Verify that the transaction data CSV file exists"""
        path = os.path.join(self.REPO_DIR, "data/transactions.csv")
        assert os.path.exists(path), f"transactions.csv not found at {path}"

    def test_build_script_is_valid_python(self):
        """Verify that the build script has valid Python syntax"""
        import ast
        path = os.path.join(self.REPO_DIR, "scripts/build_financial_report.py")
        with open(path, "r") as f:
            source = f.read()
        try:
            ast.parse(source)
        except SyntaxError as e:
            pytest.fail(f"build_financial_report.py has syntax error: {e}")

    # === Semantic Checks ===

    def test_transactions_csv_has_correct_headers(self):
        """Verify that transactions.csv has the required column headers"""
        path = os.path.join(self.REPO_DIR, "data/transactions.csv")
        with open(path, "r", newline="") as f:
            reader = csv.reader(f)
            headers = next(reader)

        headers_lower = [h.strip().lower() for h in headers]
        required_headers = ["date", "description", "category", "amount", "type"]
        for h in required_headers:
            assert any(h in hdr for hdr in headers_lower), (
                f"transactions.csv missing required header: {h}. "
                f"Found headers: {headers}"
            )

    def test_transactions_csv_has_200_rows(self):
        """Verify that transactions.csv has approximately 200 data rows"""
        path = os.path.join(self.REPO_DIR, "data/transactions.csv")
        with open(path, "r", newline="") as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            rows = list(reader)

        assert len(rows) >= 195, (
            f"Expected approximately 200 data rows, got {len(rows)}"
        )

    def test_transactions_csv_has_revenue_and_expense(self):
        """Verify that transactions include both revenue and expense types"""
        path = os.path.join(self.REPO_DIR, "data/transactions.csv")
        with open(path, "r", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        # Find the type column (case-insensitive)
        type_col = None
        for key in rows[0].keys():
            if key.strip().lower() == "type":
                type_col = key
                break

        assert type_col is not None, "CSV missing 'type' column"

        types = set(row[type_col].strip().lower() for row in rows)
        assert "revenue" in types, "transactions.csv should have revenue type entries"
        assert "expense" in types, "transactions.csv should have expense type entries"

    def test_build_script_uses_openpyxl(self):
        """Verify that the build script uses openpyxl library"""
        path = os.path.join(self.REPO_DIR, "scripts/build_financial_report.py")
        with open(path, "r") as f:
            content = f.read()

        assert "openpyxl" in content, (
            "build_financial_report.py should import/use openpyxl"
        )

    def test_build_script_creates_four_sheets(self):
        """Verify that the build script creates 4 worksheets"""
        path = os.path.join(self.REPO_DIR, "scripts/build_financial_report.py")
        with open(path, "r") as f:
            content = f.read()

        expected_sheets = ["Transactions", "Monthly Summary", "P&L Statement", "Dashboard"]
        found_count = sum(1 for sheet in expected_sheets if sheet in content)
        assert found_count >= 3, (
            f"Script should create 4 sheets. Found {found_count} of {len(expected_sheets)} "
            f"expected sheet names: {expected_sheets}"
        )

    def test_build_script_uses_formulas_not_hardcoded(self):
        """Verify that the build script uses Excel formulas (SUMPRODUCT, SUM, etc.)"""
        path = os.path.join(self.REPO_DIR, "scripts/build_financial_report.py")
        with open(path, "r") as f:
            content = f.read()

        formula_keywords = ["SUMPRODUCT", "SUM(", "COUNTIFS", "IF("]
        found_formulas = sum(1 for kw in formula_keywords if kw in content)
        assert found_formulas >= 2, (
            f"Script should use Excel formulas (SUMPRODUCT, SUM, COUNTIFS). "
            f"Found {found_formulas} formula references"
        )

    def test_build_script_creates_charts(self):
        """Verify that the build script creates charts using openpyxl"""
        path = os.path.join(self.REPO_DIR, "scripts/build_financial_report.py")
        with open(path, "r") as f:
            content = f.read()

        chart_types = ["BarChart", "LineChart", "PieChart"]
        found_charts = sum(1 for ct in chart_types if ct in content)
        assert found_charts >= 2, (
            f"Script should create multiple chart types. "
            f"Found {found_charts} of {len(chart_types)} expected chart types"
        )

    # === Functional Checks ===

    def test_build_script_runs_successfully(self):
        """Verify that the build script runs and generates the output file"""
        # Ensure openpyxl is available
        install_result = subprocess.run(
            ["pip", "install", "openpyxl"],
            capture_output=True,
            text=True,
            timeout=60,
        )

        result = subprocess.run(
            ["python", "scripts/build_financial_report.py"],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=120,
        )
        assert result.returncode == 0, (
            f"Build script failed:\n{result.stdout}\n{result.stderr}"
        )

    def test_output_xlsx_file_created(self):
        """Verify that the output Excel file is created after running the script"""
        # Run the script first
        subprocess.run(
            ["pip", "install", "openpyxl"],
            capture_output=True, text=True, timeout=60,
        )
        subprocess.run(
            ["python", "scripts/build_financial_report.py"],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120,
        )

        output_path = os.path.join(self.REPO_DIR, "output/financial_report.xlsx")
        assert os.path.exists(output_path), (
            f"Output file not found at {output_path}"
        )
        # Verify file is non-empty
        assert os.path.getsize(output_path) > 1000, (
            "Output file seems too small to be a valid Excel workbook"
        )

    def test_output_xlsx_has_four_sheets(self):
        """Verify that the output workbook has exactly 4 sheets with correct names"""
        subprocess.run(
            ["pip", "install", "openpyxl"],
            capture_output=True, text=True, timeout=60,
        )
        subprocess.run(
            ["python", "scripts/build_financial_report.py"],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120,
        )

        result = subprocess.run(
            [
                "python", "-c",
                (
                    "import openpyxl; "
                    "wb = openpyxl.load_workbook('output/financial_report.xlsx'); "
                    "sheets = wb.sheetnames; "
                    "print(len(sheets)); "
                    "print('|'.join(sheets))"
                ),
            ],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"Failed to load workbook: {result.stderr}"
        lines = result.stdout.strip().split("\n")
        sheet_count = int(lines[0])
        assert sheet_count >= 4, f"Expected at least 4 sheets, got {sheet_count}"

        sheet_names = lines[1].lower()
        assert "transaction" in sheet_names, "Missing Transactions sheet"
        assert "monthly" in sheet_names or "summary" in sheet_names, "Missing Monthly Summary sheet"

    def test_transactions_sheet_has_data(self):
        """Verify that the Transactions sheet has formatted data with correct row count"""
        subprocess.run(
            ["pip", "install", "openpyxl"],
            capture_output=True, text=True, timeout=60,
        )
        subprocess.run(
            ["python", "scripts/build_financial_report.py"],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120,
        )

        result = subprocess.run(
            [
                "python", "-c",
                (
                    "import openpyxl; "
                    "wb = openpyxl.load_workbook('output/financial_report.xlsx'); "
                    "ws = wb.worksheets[0]; "
                    "print(ws.max_row); "
                    "print(ws.cell(1,1).value)"
                ),
            ],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"Failed to read worksheet: {result.stderr}"
        lines = result.stdout.strip().split("\n")
        row_count = int(lines[0])
        assert row_count >= 200, (
            f"Transactions sheet should have at least 200 data rows, got {row_count}"
        )

    def test_monthly_summary_has_formulas(self):
        """Verify that the Monthly Summary sheet uses Excel formulas, not hardcoded values"""
        subprocess.run(
            ["pip", "install", "openpyxl"],
            capture_output=True, text=True, timeout=60,
        )
        subprocess.run(
            ["python", "scripts/build_financial_report.py"],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120,
        )

        result = subprocess.run(
            [
                "python", "-c",
                (
                    "import openpyxl; "
                    "wb = openpyxl.load_workbook('output/financial_report.xlsx'); "
                    "# Find Monthly Summary sheet\n"
                    "ws = None; "
                    "for s in wb.worksheets: "
                    "    if 'monthly' in s.title.lower() or 'summary' in s.title.lower(): "
                    "        ws = s; break\n"
                    "if ws is None: ws = wb.worksheets[1]\n"
                    "formulas = []; "
                    "for row in ws.iter_rows(min_row=2, max_row=14, max_col=6): "
                    "    for cell in row: "
                    "        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='): "
                    "            formulas.append(cell.value)\n"
                    "print(len(formulas)); "
                    "if formulas: print(formulas[0][:50])"
                ),
            ],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"Failed to read formulas: {result.stderr}"
        lines = result.stdout.strip().split("\n")
        formula_count = int(lines[0])
        assert formula_count >= 10, (
            f"Monthly Summary should have formulas, not hardcoded values. "
            f"Found {formula_count} formula cells"
        )

    def test_dashboard_has_charts(self):
        """Verify that the Dashboard sheet contains charts"""
        subprocess.run(
            ["pip", "install", "openpyxl"],
            capture_output=True, text=True, timeout=60,
        )
        subprocess.run(
            ["python", "scripts/build_financial_report.py"],
            cwd=self.REPO_DIR,
            capture_output=True, text=True, timeout=120,
        )

        result = subprocess.run(
            [
                "python", "-c",
                (
                    "import openpyxl; "
                    "wb = openpyxl.load_workbook('output/financial_report.xlsx'); "
                    "# Find Dashboard sheet\n"
                    "ws = None; "
                    "for s in wb.worksheets: "
                    "    if 'dashboard' in s.title.lower(): "
                    "        ws = s; break\n"
                    "if ws is None: ws = wb.worksheets[-1]\n"
                    "chart_count = len(ws._charts); "
                    "print(chart_count)"
                ),
            ],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"Failed to read charts: {result.stderr}"
        chart_count = int(result.stdout.strip())
        assert chart_count >= 3, (
            f"Dashboard should have at least 3 charts, found {chart_count}"
        )
