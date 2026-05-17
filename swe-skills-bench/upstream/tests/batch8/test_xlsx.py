"""
Test skill: xlsx
Verify that the Agent correctly implements a financial report generator
using openpyxl with multi-sheet workbook, formulas, formatting, charts,
data validation, and a sensitivity analysis table.
"""

import os
import subprocess
import sys
import ast
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"

    # === File Path Checks ===

    def test_financial_report_script_exists(self):
        """Verify that the main financial report script exists"""
        filepath = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        assert os.path.exists(filepath), f"financial_report.py not found at {filepath}"

    def test_formatting_module_exists(self):
        """Verify that the formatting utilities module exists"""
        filepath = os.path.join(self.REPO_DIR, "openpyxl/sample/formatting.py")
        assert os.path.exists(filepath), f"formatting.py not found at {filepath}"

    def test_formulas_module_exists(self):
        """Verify that the formula builder module exists"""
        filepath = os.path.join(self.REPO_DIR, "openpyxl/sample/formulas.py")
        assert os.path.exists(filepath), f"formulas.py not found at {filepath}"

    def test_charts_module_exists(self):
        """Verify that the charts module exists"""
        filepath = os.path.join(self.REPO_DIR, "openpyxl/sample/charts.py")
        assert os.path.exists(filepath), f"charts.py not found at {filepath}"

    def test_all_modules_valid_python(self):
        """Verify all sample modules are valid Python syntax"""
        modules = [
            "openpyxl/sample/financial_report.py",
            "openpyxl/sample/formatting.py",
            "openpyxl/sample/formulas.py",
            "openpyxl/sample/charts.py",
        ]
        for mod in modules:
            filepath = os.path.join(self.REPO_DIR, mod)
            if os.path.exists(filepath):
                with open(filepath) as f:
                    content = f.read()
                try:
                    ast.parse(content)
                except SyntaxError as e:
                    pytest.fail(f"{mod} has syntax errors: {e}")

    # === Semantic Checks ===

    def test_financial_report_creates_four_sheets(self):
        """Verify that the script references all 4 required sheet names"""
        filepath = os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.py")
        with open(filepath) as f:
            content = f.read()

        required_sheets = ["Assumptions", "Income Statement", "Balance Sheet", "Dashboard"]
        for sheet in required_sheets:
            assert sheet in content, (
                f"financial_report.py missing reference to sheet '{sheet}'"
            )

    def test_formulas_module_has_excel_formulas(self):
        """Verify formulas.py contains Excel formula strings, not hardcoded values"""
        filepath = os.path.join(self.REPO_DIR, "openpyxl/sample/formulas.py")
        with open(filepath) as f:
            content = f.read()

        # Excel formulas start with = and reference cells
        has_formula_patterns = (
            "=" in content and (
                "Assumptions!" in content
                or "!" in content  # Cross-sheet references
                or re.search(r'[A-Z]+\d+', content) is not None  # Cell references like A1, B2
            )
        )
        assert has_formula_patterns, (
            "formulas.py does not appear to contain Excel formula strings. "
            "Computed cells must use Excel formulas, not Python-computed values."
        )

    def test_charts_module_creates_charts(self):
        """Verify charts.py creates revenue and margin charts"""
        filepath = os.path.join(self.REPO_DIR, "openpyxl/sample/charts.py")
        with open(filepath) as f:
            content = f.read()

        has_chart_creation = (
            "BarChart" in content or "LineChart" in content
            or "barchart" in content.lower() or "linechart" in content.lower()
            or "Chart" in content
        )
        assert has_chart_creation, (
            "charts.py does not contain chart creation code. "
            "Expected BarChart for revenue and LineChart for margins."
        )

    def test_formatting_has_color_coding(self):
        """Verify formatting.py implements color coding (blue for inputs, black for formulas)"""
        filepath = os.path.join(self.REPO_DIR, "openpyxl/sample/formatting.py")
        with open(filepath) as f:
            content = f.read()

        has_colors = (
            "0000FF" in content or "blue" in content.lower()  # Blue for inputs
            or "FF0000" in content or "red" in content.lower()  # Red for negatives
            or "008000" in content or "green" in content.lower()  # Green for cross-refs
        )
        assert has_colors, (
            "formatting.py does not implement color coding. "
            "Expected blue (#0000FF) for inputs, red for negatives, green for cross-refs."
        )

    # === Functional Checks ===

    def test_generate_financial_report(self):
        """Verify that running the financial report script generates an xlsx file"""
        # Install openpyxl if needed
        subprocess.run(
            ["pip", "install", "-e", "."],
            cwd=self.REPO_DIR,
            capture_output=True,
            timeout=120
        )

        result = subprocess.run(
            ["python", "openpyxl/sample/financial_report.py"],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=120
        )
        assert result.returncode == 0, (
            f"financial_report.py failed: {result.stderr[:2000]}"
        )

        # Check that an xlsx file was generated
        possible_paths = [
            os.path.join(self.REPO_DIR, "financial_report.xlsx"),
            os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.xlsx"),
        ]
        xlsx_exists = any(os.path.exists(p) for p in possible_paths)
        assert xlsx_exists, (
            "No financial_report.xlsx generated. Expected output file."
        )

    def test_workbook_has_four_sheets(self):
        """Verify the generated workbook contains exactly 4 sheets with correct names"""
        try:
            import openpyxl
        except ImportError:
            subprocess.run(["pip", "install", "-e", "."], cwd=self.REPO_DIR, capture_output=True, timeout=120)
            import openpyxl

        # Generate the workbook first
        subprocess.run(
            ["python", "openpyxl/sample/financial_report.py"],
            cwd=self.REPO_DIR, capture_output=True, timeout=120
        )

        possible_paths = [
            os.path.join(self.REPO_DIR, "financial_report.xlsx"),
            os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.xlsx"),
        ]
        xlsx_path = next((p for p in possible_paths if os.path.exists(p)), None)
        if xlsx_path is None:
            pytest.skip("financial_report.xlsx not found")

        wb = openpyxl.load_workbook(xlsx_path)
        sheet_names = wb.sheetnames
        assert len(sheet_names) == 4, (
            f"Expected 4 sheets, found {len(sheet_names)}: {sheet_names}"
        )
        expected = {"Assumptions", "Income Statement", "Balance Sheet", "Dashboard"}
        actual = set(sheet_names)
        assert expected == actual, (
            f"Expected sheets {expected}, found {actual}"
        )

    def test_workbook_has_excel_formulas(self):
        """Verify that computed cells use Excel formulas, not static values"""
        try:
            import openpyxl
        except ImportError:
            subprocess.run(["pip", "install", "-e", "."], cwd=self.REPO_DIR, capture_output=True, timeout=120)
            import openpyxl

        subprocess.run(
            ["python", "openpyxl/sample/financial_report.py"],
            cwd=self.REPO_DIR, capture_output=True, timeout=120
        )

        possible_paths = [
            os.path.join(self.REPO_DIR, "financial_report.xlsx"),
            os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.xlsx"),
        ]
        xlsx_path = next((p for p in possible_paths if os.path.exists(p)), None)
        if xlsx_path is None:
            pytest.skip("financial_report.xlsx not found")

        wb = openpyxl.load_workbook(xlsx_path)

        # Check Income Statement sheet for formulas
        formula_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1

        assert formula_count >= 10, (
            f"Expected at least 10 Excel formulas in the workbook, found {formula_count}. "
            "Computed cells must use Excel formulas, not Python-computed static values."
        )

    def test_workbook_has_data_validation(self):
        """Verify that the Assumptions sheet has data validation rules"""
        try:
            import openpyxl
        except ImportError:
            subprocess.run(["pip", "install", "-e", "."], cwd=self.REPO_DIR, capture_output=True, timeout=120)
            import openpyxl

        subprocess.run(
            ["python", "openpyxl/sample/financial_report.py"],
            cwd=self.REPO_DIR, capture_output=True, timeout=120
        )

        possible_paths = [
            os.path.join(self.REPO_DIR, "financial_report.xlsx"),
            os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.xlsx"),
        ]
        xlsx_path = next((p for p in possible_paths if os.path.exists(p)), None)
        if xlsx_path is None:
            pytest.skip("financial_report.xlsx not found")

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Assumptions"]
        validations = ws.data_validations.dataValidation if ws.data_validations else []
        assert len(validations) >= 3, (
            f"Expected at least 3 data validation rules on Assumptions sheet, found {len(validations)}. "
            "Growth rate, COGS %, SG&A %, and tax rate should have validation ranges."
        )

    def test_workbook_has_charts(self):
        """Verify that the Dashboard sheet contains charts"""
        try:
            import openpyxl
        except ImportError:
            subprocess.run(["pip", "install", "-e", "."], cwd=self.REPO_DIR, capture_output=True, timeout=120)
            import openpyxl

        subprocess.run(
            ["python", "openpyxl/sample/financial_report.py"],
            cwd=self.REPO_DIR, capture_output=True, timeout=120
        )

        possible_paths = [
            os.path.join(self.REPO_DIR, "financial_report.xlsx"),
            os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.xlsx"),
        ]
        xlsx_path = next((p for p in possible_paths if os.path.exists(p)), None)
        if xlsx_path is None:
            pytest.skip("financial_report.xlsx not found")

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Dashboard"]
        charts = ws._charts
        assert len(charts) >= 2, (
            f"Expected at least 2 charts on Dashboard sheet, found {len(charts)}. "
            "Need revenue bar chart and margin line chart."
        )
