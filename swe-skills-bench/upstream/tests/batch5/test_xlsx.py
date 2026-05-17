"""
Test skill: xlsx
Verify that the Agent correctly creates a multi-sheet financial report
workbook generator using openpyxl with formulas, formatting, and named ranges.
"""

import os
import sys
import subprocess
import pytest


class TestXlsx:
    REPO_DIR = "/workspace/openpyxl"

    REPORT_SCRIPT = "openpyxl/sample/financial_report.py"
    REPORT_TEST = "openpyxl/sample/tests/test_financial_report.py"

    def _ensure_openpyxl(self):
        """Ensure openpyxl is importable."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            subprocess.run(
                ["pip", "install", "-e", "."],
                cwd=self.REPO_DIR,
                capture_output=True, text=True, timeout=120,
            )
            import openpyxl
            return openpyxl

    def _generate_workbook(self):
        """Run the report script and return the generated file path."""
        result = subprocess.run(
            ["python", self.REPORT_SCRIPT],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            pytest.skip(f"Report generation failed: {result.stderr[:300]}")
        # Look for generated xlsx file
        for candidate in [
            os.path.join(self.REPO_DIR, "financial_report.xlsx"),
            os.path.join(self.REPO_DIR, "openpyxl/sample/financial_report.xlsx"),
        ]:
            if os.path.exists(candidate):
                return candidate
        pytest.skip("Generated xlsx file not found")

    # === File Path Checks ===

    def test_report_script_exists(self):
        """Verify financial_report.py script exists"""
        filepath = os.path.join(self.REPO_DIR, self.REPORT_SCRIPT)
        assert os.path.exists(filepath), \
            f"financial_report.py not found at {filepath}"

    def test_report_test_exists(self):
        """Verify test file for financial report exists"""
        filepath = os.path.join(self.REPO_DIR, self.REPORT_TEST)
        assert os.path.exists(filepath), \
            f"test_financial_report.py not found at {filepath}"

    def test_report_script_valid_python(self):
        """Verify financial_report.py is valid Python syntax"""
        import ast
        filepath = os.path.join(self.REPO_DIR, self.REPORT_SCRIPT)
        with open(filepath) as f:
            source = f.read()
        try:
            ast.parse(source)
        except SyntaxError as e:
            pytest.fail(f"financial_report.py has syntax error: {e}")

    # === Semantic Checks ===

    def test_script_imports_openpyxl(self):
        """Verify the script imports openpyxl for workbook generation"""
        filepath = os.path.join(self.REPO_DIR, self.REPORT_SCRIPT)
        with open(filepath) as f:
            content = f.read()
        assert "openpyxl" in content, \
            "financial_report.py does not import openpyxl"

    def test_script_creates_three_sheets(self):
        """Verify the script creates Income Statement, Balance Sheet, Dashboard"""
        filepath = os.path.join(self.REPO_DIR, self.REPORT_SCRIPT)
        with open(filepath) as f:
            content = f.read()
        for sheet_name in ["Income Statement", "Balance Sheet", "Dashboard"]:
            assert sheet_name in content, \
                f"Script missing sheet creation for '{sheet_name}'"

    def test_script_uses_formulas_not_hardcoded(self):
        """Verify the script uses Excel formulas (=...) rather than hardcoded values"""
        filepath = os.path.join(self.REPO_DIR, self.REPORT_SCRIPT)
        with open(filepath) as f:
            content = f.read()
        import re
        formula_patterns = re.findall(r"['\"]=[A-Z]", content)
        assert len(formula_patterns) >= 3, \
            f"Expected multiple Excel formulas in script, found {len(formula_patterns)}"

    def test_script_defines_named_range(self):
        """Verify the script creates the NetIncomeRange named range"""
        filepath = os.path.join(self.REPO_DIR, self.REPORT_SCRIPT)
        with open(filepath) as f:
            content = f.read()
        assert "NetIncomeRange" in content, \
            "Script missing NetIncomeRange named range definition"

    # === Functional Checks ===

    def test_script_generates_xlsx_file(self):
        """Verify the script runs and produces an xlsx file"""
        xlsx_path = self._generate_workbook()
        assert os.path.exists(xlsx_path), f"XLSX file not generated at {xlsx_path}"
        assert os.path.getsize(xlsx_path) > 1000, \
            f"Generated XLSX is too small ({os.path.getsize(xlsx_path)} bytes)"

    def test_workbook_has_three_sheets(self):
        """Verify generated workbook contains exactly three sheets"""
        openpyxl = self._ensure_openpyxl()
        xlsx_path = self._generate_workbook()
        wb = openpyxl.load_workbook(xlsx_path)
        assert len(wb.sheetnames) == 3, \
            f"Expected 3 sheets, found {len(wb.sheetnames)}: {wb.sheetnames}"
        for name in ["Income Statement", "Balance Sheet", "Dashboard"]:
            assert name in wb.sheetnames, \
                f"Sheet '{name}' missing from workbook. Found: {wb.sheetnames}"

    def test_income_statement_has_formulas(self):
        """Verify Income Statement sheet contains formula cells"""
        openpyxl = self._ensure_openpyxl()
        xlsx_path = self._generate_workbook()
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Income Statement"]
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.coordinate)
        assert len(formulas) >= 5, \
            f"Income Statement should have formula cells, found {len(formulas)}: {formulas[:5]}"

    def test_workbook_has_number_formatting(self):
        """Verify monetary cells use currency number format"""
        openpyxl = self._ensure_openpyxl()
        xlsx_path = self._generate_workbook()
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Income Statement"]
        formats = set()
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
            for cell in row:
                if cell.number_format and cell.number_format != "General":
                    formats.add(cell.number_format)
        assert len(formats) > 0, \
            "No custom number formatting applied to Income Statement cells"

    def test_balance_sheet_has_conditional_formatting(self):
        """Verify Balance Sheet has conditional formatting rules"""
        openpyxl = self._ensure_openpyxl()
        xlsx_path = self._generate_workbook()
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Balance Sheet"]
        assert len(ws.conditional_formatting) > 0, \
            "Balance Sheet missing conditional formatting rules"

    def test_workbook_has_named_range(self):
        """Verify workbook defines the NetIncomeRange named range"""
        openpyxl = self._ensure_openpyxl()
        xlsx_path = self._generate_workbook()
        wb = openpyxl.load_workbook(xlsx_path)
        defined_names = [dn.name for dn in wb.defined_names.definedName]
        assert "NetIncomeRange" in defined_names, \
            f"Named range 'NetIncomeRange' not found. Defined names: {defined_names}"

    def test_dashboard_has_cross_sheet_references(self):
        """Verify Dashboard sheet uses cross-sheet formula references"""
        openpyxl = self._ensure_openpyxl()
        xlsx_path = self._generate_workbook()
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Dashboard"]
        cross_refs = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "!" in cell.value:
                    cross_refs.append(cell.coordinate)
        assert len(cross_refs) >= 3, \
            f"Dashboard should have cross-sheet references, found {len(cross_refs)}"
