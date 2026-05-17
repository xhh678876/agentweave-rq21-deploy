"""
Test for 'xlsx' skill — Excel & Spreadsheet Automation
Validates that the Agent implemented generate_sales_report() in
openpyxl/utils/report_engine.py with summary formulas, conditional formatting,
and trend charts.
"""

import os
import sys
import ast
import subprocess
import tempfile
import pytest


class TestXlsx:
    """Verify report_engine.py implementation for openpyxl."""

    REPO_DIR = "/workspace/openpyxl"

    @classmethod
    def setup_class(cls):
        if cls.REPO_DIR not in sys.path:
            sys.path.insert(0, cls.REPO_DIR)

    # ------------------------------------------------------------------
    # L1: file & syntax
    # ------------------------------------------------------------------

    def test_report_engine_exists(self):
        """openpyxl/utils/report_engine.py must exist."""
        fpath = os.path.join(self.REPO_DIR, "openpyxl", "utils", "report_engine.py")
        assert os.path.isfile(fpath), "report_engine.py not found"

    def test_report_engine_compiles(self):
        """report_engine.py must compile without syntax errors."""
        result = subprocess.run(
            ["python", "-m", "py_compile", "openpyxl/utils/report_engine.py"],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"Syntax error:\n{result.stderr}"

    # ------------------------------------------------------------------
    # L2: structural verification via AST
    # ------------------------------------------------------------------

    def test_generate_sales_report_function_exists(self):
        """generate_sales_report function must be defined."""
        fpath = os.path.join(self.REPO_DIR, "openpyxl", "utils", "report_engine.py")
        with open(fpath, "r", encoding="utf-8") as f:
            tree = ast.parse(f.read())
        func_names = [
            n.name
            for n in ast.walk(tree)
            if isinstance(n, (ast.FunctionDef, ast.AsyncFunctionDef))
        ]
        assert (
            "generate_sales_report" in func_names
        ), f"generate_sales_report not found; functions: {func_names}"

    # ------------------------------------------------------------------
    # L2: runtime verification — generate and validate xlsx
    # ------------------------------------------------------------------

    def _generate_report(self, tmp_path):
        """Helper: call generate_sales_report and return the output path."""
        script = f"""
import sys
sys.path.insert(0, '{self.REPO_DIR}')
from openpyxl.utils.report_engine import generate_sales_report

data = [
    {{"month": "Jan", "product": "Widget", "amount": 1200}},
    {{"month": "Feb", "product": "Widget", "amount": 1100}},
    {{"month": "Mar", "product": "Widget", "amount": 1300}},
    {{"month": "Apr", "product": "Widget", "amount": 900}},
    {{"month": "May", "product": "Gadget", "amount": 1500}},
    {{"month": "Jun", "product": "Gadget", "amount": 1400}},
]
output = '{tmp_path}'
generate_sales_report(data, output)
print("DONE")
"""
        result = subprocess.run(
            ["python", "-c", script],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=120,
        )
        return result

    def test_generate_report_runs(self):
        """generate_sales_report must execute without errors."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            result = self._generate_report(tmp_path)
            assert result.returncode == 0, f"Report generation failed:\n{result.stderr}"
            assert "DONE" in result.stdout
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_generated_file_is_valid_xlsx(self):
        """Generated file must be loadable by openpyxl."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            gen = self._generate_report(tmp_path)
            if gen.returncode != 0:
                pytest.skip(f"Report generation failed: {gen.stderr[:500]}")

            from openpyxl import load_workbook

            wb = load_workbook(tmp_path)
            assert wb is not None
            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_report_has_three_sheets(self):
        """Generated workbook must contain at least 3 sheets."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            gen = self._generate_report(tmp_path)
            if gen.returncode != 0:
                pytest.skip(f"Report generation failed: {gen.stderr[:500]}")

            from openpyxl import load_workbook

            wb = load_workbook(tmp_path)
            assert (
                len(wb.sheetnames) >= 3
            ), f"Expected >= 3 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"
            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_sheet1_has_summary_formulas(self):
        """Sheet1 must contain SUM and/or AVERAGE formulas."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            gen = self._generate_report(tmp_path)
            if gen.returncode != 0:
                pytest.skip(f"Report generation failed: {gen.stderr[:500]}")

            from openpyxl import load_workbook

            wb = load_workbook(tmp_path)
            ws = wb.worksheets[0]
            formulas_found = []
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        formulas_found.append(val)
            wb.close()
            has_sum = any("SUM" in f.upper() for f in formulas_found)
            has_avg = any("AVERAGE" in f.upper() for f in formulas_found)
            assert (
                has_sum or has_avg
            ), f"No SUM/AVERAGE formulas found in Sheet1. Formulas: {formulas_found}"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_sheet2_has_conditional_formatting(self):
        """Sheet2 must have conditional formatting rules."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            gen = self._generate_report(tmp_path)
            if gen.returncode != 0:
                pytest.skip(f"Report generation failed: {gen.stderr[:500]}")

            from openpyxl import load_workbook

            wb = load_workbook(tmp_path)
            ws = wb.worksheets[1]
            cf_rules = ws.conditional_formatting
            assert (
                len(list(cf_rules)) >= 1
            ), "No conditional formatting rules found on Sheet2"
            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_sheet3_has_chart(self):
        """Sheet3 must contain at least one chart."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            gen = self._generate_report(tmp_path)
            if gen.returncode != 0:
                pytest.skip(f"Report generation failed: {gen.stderr[:500]}")

            from openpyxl import load_workbook

            wb = load_workbook(tmp_path)
            ws = wb.worksheets[2]
            assert len(ws._charts) >= 1, "No chart found on Sheet3"
            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_chart_is_line_chart(self):
        """Sheet3 chart should be a LineChart for trend visualization."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            gen = self._generate_report(tmp_path)
            if gen.returncode != 0:
                pytest.skip(f"Report generation failed: {gen.stderr[:500]}")

            from openpyxl import load_workbook
            from openpyxl.chart import LineChart

            wb = load_workbook(tmp_path)
            ws = wb.worksheets[2]
            line_charts = [c for c in ws._charts if isinstance(c, LineChart)]
            assert (
                len(line_charts) >= 1
            ), f"Expected a LineChart on Sheet3; chart types: {[type(c).__name__ for c in ws._charts]}"
            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_sheet1_contains_data_rows(self):
        """Sheet1 must contain the input data rows."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            gen = self._generate_report(tmp_path)
            if gen.returncode != 0:
                pytest.skip(f"Report generation failed: {gen.stderr[:500]}")

            from openpyxl import load_workbook

            wb = load_workbook(tmp_path)
            ws = wb.worksheets[0]
            # Should have at least header + 6 data rows + 1 summary = 8 rows
            row_count = ws.max_row
            assert (
                row_count >= 7
            ), f"Expected at least 7 rows (header+data+summary), got {row_count}"
            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
